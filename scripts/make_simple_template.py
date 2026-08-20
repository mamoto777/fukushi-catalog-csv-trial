"""
かんたん版ひな形(.xlsx)生成スクリプト v2。設計書 docs/design.md §5-12

実行: PYTHONUTF8=1 python scripts/make_simple_template.py
      PYTHONUTF8=1 python scripts/make_simple_template.py --fixtures

src/data/vocab.json を正として選択肢・シーン設定シートを機械転記する。
vocab.json を変更したら本スクリプトを再実行してひな形を作り直すこと。
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
VOCAB_PATH = ROOT / "src" / "data" / "vocab.json"

SIMPLE_HEADER = [
    "商品名",
    "メーカー",
    "分類",
    "価格(円)",
    "TAISコード",
    "ひとこと説明",
    "仕様",
    "困りごと1",
    "困りごと2",
    "誰が使う",
]

SCENE_SHEET_HEADER = ["場面", "困りごと"]

USER_CHOICES = ["本人が使う", "家族の介護に使う", "どちらも"]

SAMPLE_ROWS = [
    [
        "らくあゆみステッキ軽量型",
        "あおぞら福祉機器",
        "歩行補助",
        1200,
        "01234-000001",
        "軽くてにぎりやすい定番の一本杖",
        "重さ:290g\n高さ調節:71〜94cm(10段階)",
        "ふらつく・転びやすい",
        "屋外の外出が不安",
        "本人が使う",
    ],
    [
        "ささえ四点杖ワイド",
        "",
        "歩行補助",
        1800,
        "",
        "自立するから立ち上がり時も支えになる四点杖",
        "重さ:640g\n高さ調節:66〜89cm",
        "ふらつく・転びやすい",
        "支えがないと立てない",
        "本人が使う",
    ],
    [
        "みまもりセンサーライト",
        "みらいケア",
        "見守り・生活サポート",
        2000,
        "",
        "夜中の動きをやさしく知らせる見守りセンサー",
        "",
        "夜中に動き回る",
        "一人にするのが心配",
        "どちらも",
    ],
]

COLUMN_WIDTHS = {
    "A": 28,
    "B": 16,
    "C": 16,
    "D": 12,
    "E": 16,
    "F": 44,
    "G": 30,
    "H": 24,
    "I": 24,
    "J": 14,
}
SCENE_COLUMN_WIDTHS = {"A": 16, "B": 28}


def load_vocab():
    with open(VOCAB_PATH, encoding="utf-8") as f:
        vocab = json.load(f)
    genre_labels = [g["label"] for g in vocab["genres"]]
    scene_rows = [
        [s["label"], c] for s in vocab["scenes"] for c in s["concerns"]
    ]
    return genre_labels, scene_rows


def _write_header(ws, header, header_font, header_fill):
    for col_idx, title in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill


def build_workbook(genre_labels, scene_rows):
    wb = Workbook()

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    # ---- 商品リストシート(表示・先頭) ----
    ws = wb.active
    ws.title = "商品リスト"
    _write_header(ws, SIMPLE_HEADER, header_font, header_fill)

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"

    wrap_alignment = Alignment(wrap_text=True)
    for row_idx, row in enumerate(SAMPLE_ROWS, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 7:  # G列: 仕様
                cell.alignment = wrap_alignment

    # ---- シーン設定シート(表示) ----
    ws_scene = wb.create_sheet("シーン設定")
    _write_header(ws_scene, SCENE_SHEET_HEADER, header_font, header_fill)
    for col_letter, width in SCENE_COLUMN_WIDTHS.items():
        ws_scene.column_dimensions[col_letter].width = width
    for i, (scene_label, concern) in enumerate(scene_rows, start=2):
        ws_scene.cell(row=i, column=1, value=scene_label)
        ws_scene.cell(row=i, column=2, value=concern)

    # ---- 選択肢シート(非表示) ----
    ws_choices = wb.create_sheet("選択肢")
    ws_choices.cell(row=1, column=1, value="分類")
    ws_choices.cell(row=1, column=2, value="誰が使う")
    for i, label in enumerate(genre_labels, start=2):
        ws_choices.cell(row=i, column=1, value=label)
    for i, label in enumerate(USER_CHOICES, start=2):
        ws_choices.cell(row=i, column=2, value=label)
    ws_choices.sheet_state = "hidden"

    # ---- データ入力規則(プルダウン) ----
    dv_genre = DataValidation(
        type="list",
        formula1=f"'選択肢'!$A$2:$A${1 + len(genre_labels)}",
        allow_blank=True,
        showErrorMessage=True,
    )
    dv_concern = DataValidation(
        type="list",
        formula1="'シーン設定'!$B$2:$B$121",
        allow_blank=True,
        showErrorMessage=True,
    )
    dv_user = DataValidation(
        type="list",
        formula1=f"'選択肢'!$B$2:$B${1 + len(USER_CHOICES)}",
        allow_blank=True,
        showErrorMessage=True,
    )
    ws.add_data_validation(dv_genre)
    ws.add_data_validation(dv_concern)
    ws.add_data_validation(dv_user)
    dv_genre.add("C2:C1001")
    dv_concern.add("H2:H1001")
    dv_concern.add("I2:I1001")
    dv_user.add("J2:J1001")

    return wb


def build_fixture_workbook(with_scene_sheet: bool):
    """テスト用: 入力規則なし・見出し+サンプル3行のみ(with_scene_sheet=Falseならシーン設定シートなし)"""
    genre_labels, scene_rows = load_vocab()
    wb = Workbook()
    ws = wb.active
    ws.title = "商品リスト"
    for col_idx, title in enumerate(SIMPLE_HEADER, start=1):
        ws.cell(row=1, column=col_idx, value=title)
    for row_idx, row in enumerate(SAMPLE_ROWS, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    if with_scene_sheet:
        ws_scene = wb.create_sheet("シーン設定")
        for col_idx, title in enumerate(SCENE_SHEET_HEADER, start=1):
            ws_scene.cell(row=1, column=col_idx, value=title)
        for i, (scene_label, concern) in enumerate(scene_rows, start=2):
            ws_scene.cell(row=i, column=1, value=scene_label)
            ws_scene.cell(row=i, column=2, value=concern)

    return wb


def main():
    genre_labels, scene_rows = load_vocab()

    print(f"ジャンル: {len(genre_labels)}件")
    print(f"困りごと: {len(scene_rows)}件")
    print(f"誰が使う: {len(USER_CHOICES)}件")
    print(f"シーン設定転記: {len(scene_rows)}行")

    if "--fixtures" in sys.argv:
        out_dir = ROOT / "tests" / "fixtures"
        out_dir.mkdir(parents=True, exist_ok=True)

        ok_path = out_dir / "simple-ok.xlsx"
        build_fixture_workbook(with_scene_sheet=True).save(ok_path)
        print(f"生成: {ok_path}")

        noscene_path = out_dir / "simple-noscene.xlsx"
        build_fixture_workbook(with_scene_sheet=False).save(noscene_path)
        print(f"生成: {noscene_path}")
    else:
        out_path = ROOT / "src" / "assets" / "products-template-simple.xlsx"
        wb = build_workbook(genre_labels, scene_rows)
        wb.save(out_path)
        print(f"生成: {out_path}")


if __name__ == "__main__":
    main()
